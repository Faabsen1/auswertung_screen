import os
import json
import cv2
import numpy as np
from PIL import ImageGrab
import pytesseract
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, simpledialog

# Tesseract-Pfad anpassen!
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\f.morfeld\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789+-'

POSITION_FILE = "fensterkonfiguration.json"
EXCEL_FILE = "werte.xlsx"

# --- Konfigurationsabfrage direkt am Start ---
def frage_konfiguration():
    if os.path.exists(POSITION_FILE):
        root = tk.Tk()
        root.withdraw()
        antwort = messagebox.askyesno(
            "Konfiguration",
            "Haben sich Fensterposition oder Wertebereiche verändert?"
        )
        root.destroy()
        if not antwort:
            with open(POSITION_FILE, "r") as f:
                return json.load(f)
    fenster_position = select_window_area()
    root = tk.Tk()
    root.withdraw()
    anzahl_werte = simpledialog.askinteger(
        "Anzahl Werte", "Wie viele Werte sollen gelesen werden?", minvalue=1
    )
    root.destroy()
    werte_bereiche = select_value_areas(anzahl_werte, fenster_position)
    konfig = {
        "fenster_position": fenster_position,
        "anzahl_werte": anzahl_werte,
        "werte_bereiche": werte_bereiche
    }
    with open(POSITION_FILE, "w") as f:
        json.dump(konfig, f)
    return konfig

# --- Fensterposition markieren ---
def select_window_area():
    img_pil = ImageGrab.grab()
    screenshot_img = np.array(img_pil.convert('RGB'))
    screenshot_img = cv2.cvtColor(screenshot_img, cv2.COLOR_RGB2BGR)
    pos = []

    def mark_area(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            pos.clear()
            pos.append((x, y))
        elif event == cv2.EVENT_LBUTTONUP:
            pos.append((x, y))
            cv2.rectangle(screenshot_img, pos[0], pos[1], (255, 0, 0), 2)
            cv2.imshow("Fensterposition wählen", screenshot_img)

    cv2.namedWindow("Fensterposition wählen")
    cv2.setMouseCallback("Fensterposition wählen", mark_area)
    cv2.imshow("Fensterposition wählen", screenshot_img)
    print(">> Markiere das Fenster mit der Maus, dann drücke 'q'")
    while True:
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cv2.destroyWindow("Fensterposition wählen")
    if len(pos) == 2:
        x1, y1 = pos[0]
        x2, y2 = pos[1]
        return [min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2)]
    else:
        raise Exception("Fensterposition nicht gewählt!")

# --- Wertebereiche im Fenster markieren ---
def select_value_areas(anzahl_werte, fenster_position):
    img_pil = ImageGrab.grab(bbox=fenster_position)
    screenshot_img = np.array(img_pil.convert('RGB'))
    screenshot_img = cv2.cvtColor(screenshot_img, cv2.COLOR_RGB2BGR)
    bereiche = []

    def mark_area(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            mark_area.start = (x, y)
        elif event == cv2.EVENT_LBUTTONUP:
            end = (x, y)
            bereiche.append([mark_area.start[0], mark_area.start[1], end[0], end[1]])
            cv2.rectangle(screenshot_img, mark_area.start, end, (0, 255, 0), 2)
            cv2.imshow("Wertebereich markieren", screenshot_img)
    mark_area.start = (0, 0)

    cv2.namedWindow("Wertebereich markieren")
    cv2.setMouseCallback("Wertebereich markieren", mark_area)
    cv2.imshow("Wertebereich markieren", screenshot_img)
    print(f">> Markiere {anzahl_werte} Wertebereiche im Fenster, jeweils mit der Maus, dann drücke 'q'")
    while len(bereiche) < anzahl_werte:
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cv2.destroyWindow("Wertebereich markieren")
    return bereiche

# --- Excel vorbereiten ---
def prepare_excel(file, anzahl_werte):
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        header = ["Zeitstempel"] + [f"Wert {i+1}" for i in range(anzahl_werte)]
        ws.append(header)
        wb.save(file)
    wb = load_workbook(file)
    ws = wb.active
    return wb, ws

# --- OCR Hilfsfunktion für Vorzeichen ---
def clean_ocr_value(val):
    val = val.replace('−', '-').replace('–', '-').replace(' ', '')
    import re
    val = re.sub(r'[^0-9\+\-]', '', val)
    return val

# --- Hauptlogik ---
if __name__ == "__main__":
    konfig = frage_konfiguration()
    fenster_position = konfig["fenster_position"]
    werte_bereiche = konfig["werte_bereiche"]
    anzahl_werte = konfig["anzahl_werte"]

    wb, ws = prepare_excel(EXCEL_FILE, anzahl_werte)

    letzter_wert1 = None

    while True:
        # Screenshot vom Fensterbereich
        img_pil = ImageGrab.grab(bbox=fenster_position)
        screenshot_img = np.array(img_pil.convert('RGB'))
        screenshot_img = cv2.cvtColor(screenshot_img, cv2.COLOR_RGB2BGR)

        # Werte dynamisch auslesen
        werte = []
        for (x1, y1, x2, y2) in werte_bereiche:
            wert_img = screenshot_img[y1:y2, x1:x2]
            wert_raw = pytesseract.image_to_string(wert_img, config=config).strip()
            werte.append(clean_ocr_value(wert_raw))

        # Nur speichern, wenn sich Wert 1 geändert hat
        if letzter_wert1 is None:
            letzter_wert1 = werte[0]
        if werte[0] != letzter_wert1:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp] + werte)
            wb.save(EXCEL_FILE)
            print(f"Werte gespeichert: {werte}")
            letzter_wert1 = werte[0]

        # Kurze Pause, damit die Schleife nicht zu schnell läuft
        import time
        time.sleep(0.2)
