import pygetwindow as gw
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import mss
import numpy as np
import cv2
import pytesseract
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json
import time

# Tesseract-Pfad anpassen!
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\f.morfeld\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789+-'

POSITION_FILE = "fensterkonfiguration.json"
EXCEL_FILE = "werte.xlsx"

def select_window_from_list():
    windows = [w for w in gw.getAllWindows() if w.title and w.width > 0 and w.height > 0]
    if not windows:
        raise Exception("Keine aktiven Fenster gefunden!")
    root = tk.Tk()
    root.title("Fenster auswählen")
    tk.Label(root, text="Bitte wähle ein Fenster aus:").pack(padx=10, pady=10)
    combo = ttk.Combobox(root, state="readonly", width=80)
    combo['values'] = [f"{w.title} ({w.left},{w.top},{w.width},{w.height})" for w in windows]
    combo.current(0)
    combo.pack(padx=10, pady=10)
    selected = {'index': 0}
    def on_ok():
        selected['index'] = combo.current()
        root.destroy()
    tk.Button(root, text="OK", command=on_ok).pack(pady=10)
    root.mainloop()
    win = windows[selected['index']]
    # Rückgabe: [x1, y1, x2, y2]
    return [win.left, win.top, win.left + win.width, win.top + win.height]

def frage_konfiguration():
    if os.path.exists(POSITION_FILE):
        root = tk.Tk()
        root.withdraw()
        antwort = messagebox.askyesno(
            "Konfiguration",
            "Haben sich Fensterposition oder Wertebereiche verändert?"
        )
        root.destroy()
        if not antwort:
            with open(POSITION_FILE, "r") as f:
                return json.load(f)
    fenster_position = select_window_from_list()
    root = tk.Tk()
    root.withdraw()
    anzahl_werte = simpledialog.askinteger(
        "Anzahl Werte", "Wie viele Werte sollen gelesen werden?", minvalue=1
    )
    root.destroy()
    werte_bereiche = select_value_areas(anzahl_werte, fenster_position)
    konfig = {
        "fenster_position": fenster_position,
        "anzahl_werte": anzahl_werte,
        "werte_bereiche": werte_bereiche
    }
    with open(POSITION_FILE, "w") as f:
        json.dump(konfig, f)
    return konfig

def select_value_areas(anzahl_werte, fenster_position):
    img = screenshot_window_area(fenster_position)
    screenshot_img = img.copy()
    bereiche = []

    def mark_area(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            mark_area.start = (x, y)
        elif event == cv2.EVENT_LBUTTONUP:
            end = (x, y)
            bereiche.append([mark_area.start[0], mark_area.start[1], end[0], end[1]])
            cv2.rectangle(screenshot_img, mark_area.start, end, (0, 255, 0), 2)
            cv2.imshow("Wertebereich markieren", screenshot_img)
    mark_area.start = (0, 0)

    cv2.namedWindow("Wertebereich markieren")
    cv2.setMouseCallback("Wertebereich markieren", mark_area)
    cv2.imshow("Wertebereich markieren", screenshot_img)
    print(f">> Markiere {anzahl_werte} Wertebereiche im Fenster, jeweils mit der Maus, dann drücke 'q'")
    while len(bereiche) < anzahl_werte:
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cv2.destroyWindow("Wertebereich markieren")
    return bereiche

def screenshot_window_area(area):
    time.sleep(0.2)  # Kurze Pause, um sicherzustellen, dass das Fenster bereit ist
    left, top, right, bottom = area
    width = right - left
    height = bottom - top
    with mss.mss() as sct:
        monitor = {"left": int(left), "top": int(top), "width": int(width), "height": int(height)}
        img = np.array(sct.grab(monitor))
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
        return img

def prepare_excel(file, anzahl_werte):
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        header = ["Zeitstempel"] + [f"Wert {i+1}" for i in range(anzahl_werte)]
        ws.append(header)
        wb.save(file)
    wb = load_workbook(file)
    ws = wb.active
    return wb, ws

def normalize_value(val):
    # Entfernt Leerzeichen und wandelt Unicode-Minus in ASCII-Minus um
    return val.replace('−', '-').replace('–', '-').replace(' ', '').strip()

if __name__ == "__main__":
    konfig = frage_konfiguration()
    fenster_position = konfig["fenster_position"]
    werte_bereiche = konfig["werte_bereiche"]
    anzahl_werte = konfig["anzahl_werte"]

    wb, ws = prepare_excel(EXCEL_FILE, anzahl_werte)
    letzter_wert1 = None

    while True:
        img = screenshot_window_area(fenster_position)
        werte = []
        for (x1, y1, x2, y2) in werte_bereiche:
            wert_img = img[y1:y2, x1:x2]
            wert_raw = pytesseract.image_to_string(wert_img, config=config)
            werte.append(normalize_value(wert_raw))

        # Vergleich mit den letzten gespeicherten Werten
        if 'letzte_gespeicherte_werte' not in locals():
            letzte_gespeicherte_werte = werte.copy()

        if werte != letzte_gespeicherte_werte:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp] + werte)
            wb.save(EXCEL_FILE)
            print(f"Werte gespeichert: {werte}")
            letzte_gespeicherte_werte = werte.copy()
        
