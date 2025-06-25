import os
import json
import time
from datetime import datetime
import cv2
import mss
import numpy as np
import pygetwindow as gw
import pytesseract
from openpyxl import Workbook, load_workbook
from screeninfo import get_monitors
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import threading
import pyautogui

# Tesseract-Pfad anpassen!
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\f.morfeld\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789+-,.'

POSITION_FILE = "fensterkonfiguration.json"
EXCEL_FILE = "werte.xlsx"

def select_window_or_screen_from_list():
    """Fenster oder Monitor auswählen."""
    windows = [w for w in gw.getAllWindows() if w.title and w.width > 0 and w.height > 0]
    monitors = get_monitors()
    entries = [f"Monitor {i+1}: {m.width}x{m.height} @ ({m.x},{m.y})" for i, m in enumerate(monitors)]
    entries += [f"Fenster: {w.title} ({w.left},{w.top},{w.width},{w.height})" for w in windows]
    root = tk.Tk()
    root.title("Fenster oder Bildschirm auswählen")
    root.attributes("-topmost", True)
    tk.Label(root, text="Bitte wähle ein Fenster oder einen Bildschirm aus:").pack(padx=10, pady=10)
    combo = ttk.Combobox(root, state="readonly", width=80)
    combo['values'] = entries
    combo.current(0)
    combo.pack(padx=10, pady=10)
    selected = {'index': 0}
    def on_ok():
        selected['index'] = combo.current()
        root.destroy()
    tk.Button(root, text="OK", command=on_ok).pack(pady=10)
    root.mainloop()
    idx = selected['index']
    if idx < len(monitors):
        m = monitors[idx]
        return [m.x, m.y, m.x + m.width, m.y + m.height]
    else:
        w = windows[idx - len(monitors)]
        return [w.left, w.top, w.left + w.width, w.top + w.height]

def frage_konfiguration():
    """Frage Konfiguration ab oder lade sie aus Datei."""
    default_anzahl_werte = 1
    default_trigger_seconds = 5

    if os.path.exists(POSITION_FILE):
        with open(POSITION_FILE, "r") as f:
            last_config = json.load(f)
        default_anzahl_werte = last_config.get("anzahl_werte", 1)
        default_trigger_seconds = last_config.get("trigger_seconds", 5)

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        antwort = messagebox.askyesno(
            "Konfiguration",
            "Haben sich Fensterposition oder Wertebereiche verändert?"
        )
        root.destroy()
        if not antwort:
            return last_config
    else:
        last_config = {}

    fenster_position = select_window_or_screen_from_list()

    # Gemeinsames GUI für Anzahl Werte und Trigger
    class KonfigDialog(simpledialog.Dialog):
        def body(self, master):
            tk.Label(master, text="Wie viele Werte sollen gelesen werden?").grid(row=0, sticky="w")
            self.anzahl_entry = tk.Entry(master)
            self.anzahl_entry.insert(0, str(default_anzahl_werte))
            self.anzahl_entry.grid(row=0, column=1)

            tk.Label(master, text="Wie viele Sekunden zwischen den Messungen?").grid(row=1, sticky="w")
            self.trigger_entry = tk.Entry(master)
            self.trigger_entry.insert(0, str(default_trigger_seconds))
            self.trigger_entry.grid(row=1, column=1)
            return self.anzahl_entry

        def apply(self):
            self.result = (
                int(self.anzahl_entry.get()),
                int(self.trigger_entry.get())
            )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    dialog = KonfigDialog(root, title="Messkonfiguration")
    anzahl_werte, trigger_seconds = dialog.result

    werte_bereiche = select_value_areas(anzahl_werte, fenster_position)
    konfig = {
        "fenster_position": fenster_position,
        "anzahl_werte": anzahl_werte,
        "trigger_seconds": trigger_seconds,
        "werte_bereiche": werte_bereiche
    }
    with open(POSITION_FILE, "w") as f:
        json.dump(konfig, f)
    return konfig

def screenshot_window_area(area):
    """Screenshot eines Bereichs aufnehmen."""
    time.sleep(0.2)
    left, top, right, bottom = area
    width = right - left
    height = bottom - top
    with mss.mss() as sct:
        monitor = {"left": int(left), "top": int(top), "width": int(width), "height": int(height)}
        img = np.array(sct.grab(monitor))
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
        return img

def prepare_excel(file, anzahl_werte):
    """Excel-Datei vorbereiten oder laden."""
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        header = ["Zeitstempel"] + [f"Wert {i+1}" for i in range(anzahl_werte)]
        ws.append(header)
        wb.save(file)
    wb = load_workbook(file)
    ws = wb.active
    return wb, ws

def normalize_value(val):
    """OCR-Ausgabe normalisieren."""
    return val.replace('−', '-').replace('–', '-').replace(' ', '').strip()

def select_value_areas(anzahl_werte, fenster_position):
    """Bereiche für die Werte per Maus markieren."""
    img = screenshot_window_area(fenster_position)
    screenshot_img = img.copy()
    cv2.namedWindow("Wertebereiche markieren", cv2.WINDOW_NORMAL)
    cv2.setWindowProperty("Wertebereiche markieren", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
    cv2.imshow("Wertebereiche markieren", screenshot_img)
    print(f">> Markiere jetzt {anzahl_werte} Wertebereiche und drücke jeweils 'q'")
    wertebereiche = []
    def mark_value(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            mark_value.start = (x, y)
        elif event == cv2.EVENT_LBUTTONUP:
            end = (x, y)
            wertebereiche.append([mark_value.start[0], mark_value.start[1], end[0], end[1]])
            cv2.rectangle(screenshot_img, mark_value.start, end, (0, 255, 0), 2)
            cv2.imshow("Wertebereiche markieren", screenshot_img)
    mark_value.start = (0, 0)
    cv2.setMouseCallback("Wertebereiche markieren", mark_value)
    while len(wertebereiche) < anzahl_werte:
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cv2.destroyWindow("Wertebereiche markieren")
    return wertebereiche

def preprocess_for_ocr(img):
    """Bild für OCR vorbereiten."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return thresh

class OverlayWindow(tk.Tk):
    """Kleines Overlay-Fenster für Live-Werte."""
    def __init__(self):
        super().__init__()
        self.title("Live Werte")
        self.geometry("+10+10")
        self.attributes("-topmost", True)
        self.label = tk.Label(self, text="", font=("Consolas", 18), fg="lime", bg="black", justify="left")
        self.label.pack(padx=10, pady=10)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.closed = False
        self._start_mouse_keepalive()

    def _start_mouse_keepalive(self):
        self._mouse_keepalive()
    
    def _mouse_keepalive(self):
        try:
            pyautogui.press('volumedown')
            pyautogui.press('volumeup')
        except Exception:
            pass
        self.after(60000, self._mouse_keepalive)  # alle 60 Sekunden

    def update_values(self, trigger_seconds, werte, laufzeit_str):
        text = (
            f"Triggerzeit: {trigger_seconds} s\n"
            f"Werte: {', '.join(werte)}\n"
            f"Verstrichene Zeit: {laufzeit_str}"
        )
        self.after(0, self.label.config, {'text': text})

    def on_close(self):
        self.closed = True
        self.destroy()

if __name__ == "__main__":
    konfig = frage_konfiguration()
    fenster_position = konfig["fenster_position"]
    anzahl_werte = konfig["anzahl_werte"]
    trigger_seconds = konfig["trigger_seconds"]
    werte_bereiche = konfig["werte_bereiche"]

    wb, ws = prepare_excel(EXCEL_FILE, anzahl_werte)
    start_time = time.time()

    overlay_window = OverlayWindow()

    def mess_loop():
        while not overlay_window.closed:
            loop_start = time.time()

            img = screenshot_window_area(fenster_position)
            werte = []
            for (x1, y1, x2, y2) in werte_bereiche:
                wert_img = img[y1:y2, x1:x2]
                wert_img_proc = preprocess_for_ocr(wert_img)
                wert_raw = pytesseract.image_to_string(wert_img_proc, config=config)
                werte.append(normalize_value(wert_raw))

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp] + werte)
            wb.save(EXCEL_FILE)
            print(f"Werte gespeichert: {werte}")

            laufzeit = time.time() - start_time
            h = int(laufzeit // 3600)
            m = int((laufzeit % 3600) // 60)
            s = int(laufzeit % 60)
            laufzeit_str = f"{h:02d}:{m:02d}:{s:02d}"

            overlay_window.update_values(trigger_seconds, werte, laufzeit_str)
            print(f"Laufzeit: {laufzeit_str}")

            # Restzeit bis zum nächsten Trigger schlafen
            elapsed = time.time() - loop_start
            sleep_time = max(0, trigger_seconds - elapsed)
            time.sleep(sleep_time)

    threading.Thread(target=mess_loop, daemon=True).start()
    overlay_window.mainloop()
